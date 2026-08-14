#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_report.py
================

Преобразует типовую выгрузку сессий StartExam (например,
"sessions-2026-08-14.xlsx") в утверждённую форму отчёта
("форма_предоставления_отчета.xlsx"):

    № п/п | Компания | Тестирование | Прокторинг (4 категории по длительности)

Категории Прокторинга — это диапазоны длительности сеанса (колонка "Время"
в исходном файле):
    1 категория: 00:01 - 15:00 (мин:сек, т.е. > 0 и <= 15 минут)
    2 категория: 15:01 - 30:00
    3 категория: 30:01 - 45:00
    4 категория: 45:01 - 60:00

"Тестирование" — общее количество сеансов (строк) по компании.
Сеансы с нулевой длительностью (00:00) или длительностью более 60 минут
не попадают ни в одну из 4 категорий Прокторинга (это видно и в
исходном шаблоне: сумма 4 категорий не всегда обязана совпадать
с "Тестирование" построчно, хотя в приведённом примере совпадала).
Эта логика вынесена в функцию `categorize_minutes()` — при необходимости
её легко изменить.

ИСПОЛЬЗОВАНИЕ
--------------
    python build_report.py <входной_файл.xlsx> <выходной_файл.xlsx>

Пример:
    python build_report.py sessions-2026-08-14.xlsx report-2026-08-14.xlsx

Список компаний в отчёт НЕ задаётся заранее — берутся ровно те компании,
которые встретились в загруженном файле (колонка "Центр наименование"),
отсортированные по алфавиту. Если компании в выгрузке нет — в отчёте
её тоже не будет.
"""

import sys
import argparse
from datetime import timedelta, datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Названия колонок, которые скрипт ищет в исходном файле (регистр не важен,
# сравнение идёт по точному совпадению после strip()).
COL_COMPANY = "Центр наименование"
COL_DURATION = "Время"

CATEGORY_LABELS = [
    "1 категория 00:01-15:00",
    "2 категория 15:01-30:00",
    "3 категория 30:01-45:00",
    "4 категория 45:01-60:00",
]


def duration_to_minutes(value):
    """Приводит значение колонки 'Время' к количеству минут (float)."""
    if value is None:
        return 0.0
    if isinstance(value, timedelta):
        return value.total_seconds() / 60.0
    if isinstance(value, datetime):
        # На случай, если Excel хранит время как datetime (день=1899-12-30)
        return (value.hour * 60) + value.minute + value.second / 60.0
    if isinstance(value, (int, float)):
        # Число дней (стандарт Excel для отформатированного времени)
        return float(value) * 24 * 60
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return 0.0
        parts = s.split(":")
        try:
            parts = [float(p) for p in parts]
        except ValueError:
            return 0.0
        while len(parts) < 3:
            parts.insert(0, 0.0)
        h, m, sec = parts[-3:]
        return h * 60 + m + sec / 60.0
    return 0.0


def categorize_minutes(minutes):
    """
    Возвращает индекс категории (0-3) или None, если длительность
    не попадает ни в один из диапазонов (0 минут или > 60 минут).
    """
    if minutes <= 0:
        return None
    if minutes <= 15:
        return 0
    if minutes <= 30:
        return 1
    if minutes <= 45:
        return 2
    if minutes <= 60:
        return 3
    return None


def find_header_row(ws):
    """Находит строку заголовков в исходном файле и индексы нужных колонок."""
    for row in ws.iter_rows(min_row=1, max_row=5):
        values = [c.value for c in row]
        if COL_COMPANY in values:
            header_row_idx = row[0].row
            col_map = {}
            for cell in row:
                if cell.value is not None:
                    col_map[str(cell.value).strip()] = cell.column
            return header_row_idx, col_map
    raise ValueError(
        f"Не найдена строка заголовков с колонкой '{COL_COMPANY}' "
        f"в первых 5 строках листа '{ws.title}'."
    )


def load_source(path):
    """Читает исходную выгрузку и агрегирует данные по компаниям."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_row_idx, col_map = find_header_row(ws)
    if COL_COMPANY not in col_map:
        raise ValueError(f"В файле не найдена колонка '{COL_COMPANY}'.")
    if COL_DURATION not in col_map:
        raise ValueError(f"В файле не найдена колонка '{COL_DURATION}'.")

    company_col = col_map[COL_COMPANY]
    duration_col = col_map[COL_DURATION]

    stats = {}  # company -> {"total": int, "cats": [c1, c2, c3, c4]}
    unmatched_durations = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        company_cell = row[company_col - 1]
        company = company_cell.value
        if company is None or str(company).strip() == "":
            continue
        company = str(company).strip()

        duration_cell = row[duration_col - 1]
        minutes = duration_to_minutes(duration_cell.value)
        cat = categorize_minutes(minutes)

        entry = stats.setdefault(company, {"total": 0, "cats": [0, 0, 0, 0]})
        entry["total"] += 1
        if cat is not None:
            entry["cats"][cat] += 1
        else:
            unmatched_durations += 1

    return stats, unmatched_durations


def build_output(stats, output_path):
    """Строит выходной файл в утверждённом формате."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    thin = Side(style="thin")
    dotted = Side(style="dotted")
    border_header = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_data = Border(left=dotted, right=dotted, top=dotted, bottom=dotted)
    center = Alignment(horizontal="center", vertical="center")
    center_top_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    # --- Заголовки ---
    ws.merge_cells("A1:A2")
    ws.merge_cells("B1:B2")
    ws.merge_cells("C1:C2")
    ws.merge_cells("D1:G1")

    ws["A1"] = "№ п/п"
    ws["B1"] = "Компания"
    ws["C1"] = "Тестирование"
    ws["D1"] = "Прокторинг"
    for coord in ("A1", "B1", "C1", "D1"):
        ws[coord].alignment = center
        ws[coord].border = border_header

    for col, label in zip(("D2", "E2", "F2", "G2"), CATEGORY_LABELS):
        ws[col] = label
        ws[col].alignment = center_top_wrap
        ws[col].border = border_header

    for coord in ("A2", "B2", "C2", "E1", "F1", "G1"):
        ws[coord].border = border_header

    ws.column_dimensions["A"].width = 6.5
    ws.column_dimensions["B"].width = 45.5
    ws.column_dimensions["C"].width = 15.7
    for col in ("D", "E", "F", "G"):
        ws.column_dimensions[col].width = 12.4
    ws.row_dimensions[1].height = 14.25
    ws.row_dimensions[2].height = 28.5

    # --- Список компаний: ровно то, что встретилось в выгрузке ---
    ordered_companies = sorted(stats.keys())

    row_idx = 3
    totals = {"total": 0, "cats": [0, 0, 0, 0]}
    for i, company in enumerate(ordered_companies, start=1):
        entry = stats.get(company, {"total": 0, "cats": [0, 0, 0, 0]})
        total = entry["total"]
        cats = entry["cats"]

        ws.cell(row=row_idx, column=1, value=i).alignment = center
        ws.cell(row=row_idx, column=2, value=company).alignment = left
        ws.cell(row=row_idx, column=3, value=total).alignment = center
        for j, val in enumerate(cats):
            # Пишем явные числа (включая 0) во все 4 колонки категорий.
            # В образце формы часть пустых ячеек оставлена без значения
            # (видимо, вручную), но для автоматической генерации явные
            # нули однозначнее и корректно суммируются формулами Excel.
            ws.cell(row=row_idx, column=4 + j, value=val).alignment = center

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = border_data

        totals["total"] += total
        for k in range(4):
            totals["cats"][k] += cats[k]
        row_idx += 1

    # --- Строка ИТОГО ---
    ws.cell(row=row_idx, column=2, value="ИТОГО ").alignment = left
    ws.cell(row=row_idx, column=2).font = Font(bold=True)
    ws.cell(row=row_idx, column=3, value=totals["total"]).alignment = center
    ws.cell(row=row_idx, column=3).font = Font(bold=True)
    for j in range(4):
        c = ws.cell(row=row_idx, column=4 + j, value=totals["cats"][j])
        c.alignment = center
        c.font = Font(bold=True)
    for col in range(1, 8):
        ws.cell(row=row_idx, column=col).border = border_data

    wb.save(output_path)
    return output_path, totals


def main():
    parser = argparse.ArgumentParser(
        description="Преобразует выгрузку сессий StartExam в утверждённую "
                     "форму отчёта (Тестирование / Прокторинг по категориям)."
    )
    parser.add_argument("input", help="Путь к исходному файлу выгрузки (.xlsx)")
    parser.add_argument("output", help="Путь к выходному файлу отчёта (.xlsx)")
    args = parser.parse_args()

    stats, unmatched = load_source(args.input)
    output_path, totals = build_output(stats, args.output)

    print(f"Готово: {output_path}")
    print(f"Всего сеансов (Тестирование): {totals['total']}")
    print(f"По категориям Прокторинга: {totals['cats']}")
    if unmatched:
        print(f"Сеансов вне диапазонов категорий (0 мин или > 60 мин): {unmatched}")


if __name__ == "__main__":
    main()
